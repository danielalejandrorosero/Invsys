import json
import sys
import os
import datetime

# Intentar importar pandas, pero manejar el caso si no está instalado
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Verificar si pandas está disponible
if not PANDAS_AVAILABLE:
    # Si pandas no está disponible, intentar crear un Excel básico con openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Intentar cargar los datos
        try:
            data = json.loads(sys.argv[1])
            if not data:
                raise ValueError("No hay datos para generar el Excel")
                
            # Crear un Excel básico con los datos usando solo openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Stock"
            
            # Estilos básicos
            title_font = Font(size=14, bold=True)
            header_font = Font(size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Añadir título
            ws['A1'] = "REPORTE DE STOCK"
            ws.merge_cells('A1:H1')
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Añadir fecha
            ws['A2'] = f"Fecha: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws.merge_cells('A2:H2')
            
            # Añadir encabezados
            headers = ['Producto', 'Almacén', 'Cantidad', 'Stock Mínimo', 'Stock Máximo', 'Categoría', 'Proveedor', 'Estado']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Añadir datos
            row_num = 5
            for item in data:
                # Determinar estado del stock
                quantity = int(item.get('cantidad', 0))
                min_stock = int(item.get('stock_minimo', 0))
                max_stock = int(item.get('stock_maximo', 0))
                
                if quantity <= 0:
                    status = "Sin Stock"
                elif quantity < min_stock:
                    status = "Bajo"
                elif quantity > max_stock:
                    status = "Exceso"
                else:
                    status = "Normal"
                
                # Añadir datos a las celdas
                ws.cell(row=row_num, column=1).value = item.get('producto', '')
                ws.cell(row=row_num, column=2).value = item.get('almacen', '')
                ws.cell(row=row_num, column=3).value = quantity
                ws.cell(row=row_num, column=4).value = min_stock
                ws.cell(row=row_num, column=5).value = max_stock
                ws.cell(row=row_num, column=6).value = item.get('categoria', '')
                ws.cell(row=row_num, column=7).value = item.get('proveedor', '')
                ws.cell(row=row_num, column=8).value = status
                
                row_num += 1
            
            # Ajustar ancho de columnas
            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 15
                
            # Guardar el archivo
            output_path = os.path.join(os.getcwd(), "reporte_stock.xlsx")
            wb.save(output_path)
            print(output_path, flush=True)
            sys.exit(0)
            
        except (IndexError, json.JSONDecodeError, ValueError) as e:
            # Crear un Excel con mensaje de error
            wb = Workbook()
            ws = wb.active
            ws.title = "Error"
            ws['A1'] = "Error al generar el reporte"
            ws['A2'] = "No se encontraron datos para generar el reporte."
            
            # Guardar el archivo
            output_path = os.path.join(os.getcwd(), "reporte_error.xlsx")
            wb.save(output_path)
            print(output_path, flush=True)
            sys.exit(0)
            
    except ImportError:
        # Si ni siquiera openpyxl está disponible
        print("Error: No se pueden importar los módulos necesarios (pandas, openpyxl)", file=sys.stderr)
        sys.exit(1)

# Recibir datos desde PHP
try:
    data = json.loads(sys.argv[1])
    if not data:
        raise ValueError("No hay datos para generar el Excel")
except (IndexError, json.JSONDecodeError, ValueError) as e:
    # Crear un Excel con mensaje de error
    wb = Workbook()
    ws = wb.active
    ws.title = "Error"
    ws['A1'] = "Error al generar el reporte"
    ws['A2'] = "No se encontraron datos para generar el reporte."
    
    # Guardar el archivo
    output_path = os.path.join(os.getcwd(), "reporte_error.xlsx")
    wb.save(output_path)
    print(output_path, flush=True)
    sys.exit(0)

# Convertir los datos a un DataFrame de pandas
df = pd.DataFrame(data)

# Crear un nuevo archivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Reporte de Stock"

# Estilos para el título
title_font = Font(name='Calibri', size=16, bold=True)
subtitle_font = Font(name='Calibri', size=10)
header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

# Estilos para los encabezados
header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
align_center = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# Añadir título
ws['A1'] = "REPORTE DE STOCK"
ws.merge_cells('A1:H1')
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

# Añadir fecha
ws['A2'] = f"Fecha: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws.merge_cells('A2:H2')
ws['A2'].font = subtitle_font
ws['A2'].alignment = Alignment(horizontal='right')

# Añadir encabezados
headers = ['Producto', 'Almacén', 'Cantidad', 'Stock Mínimo', 'Stock Máximo', 'Categoría', 'Proveedor', 'Estado']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border

# Mapeo de columnas del dataframe a columnas de Excel
df_columns = ['producto', 'almacen', 'cantidad', 'stock_minimo', 'stock_maximo', 'categoria', 'proveedor']

# Añadir datos
row_num = 5
for _, row_data in df.iterrows():
    # Determinar estado del stock
    quantity = int(row_data['cantidad'])
    min_stock = int(row_data['stock_minimo'])
    max_stock = int(row_data['stock_maximo'])
    
    if quantity <= 0:
        status = "Sin Stock"
        status_color = "F44336"  # Rojo
    elif quantity < min_stock:
        status = "Bajo"
        status_color = "FF9800"  # Naranja
    elif quantity > max_stock:
        status = "Exceso"
        status_color = "2196F3"  # Azul
    else:
        status = "Normal"
        status_color = "4CAF50"  # Verde
    
    # Añadir datos a las celdas
    for col_num, column in enumerate(df_columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = row_data[column]
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Añadir celda de estado
    status_cell = ws.cell(row=row_num, column=8)
    status_cell.value = status
    status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
    status_cell.font = Font(color="FFFFFF")  # Texto blanco para mejor contraste
    status_cell.alignment = Alignment(horizontal='center')
    status_cell.border = thin_border
    
    row_num += 1

# Autoajustar ancho de columnas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Guardar el archivo Excel en una ubicación temporal con permisos adecuados
try:
    # Crear directorio temporal si no existe
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    
    # Generar nombre de archivo único
    file_name = f"reporte_stock_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(temp_dir, file_name)
    
    # Guardar el archivo
    wb.save(output_path)
    print(output_path, flush=True)
except Exception as e:
    error_msg = f"Error al guardar el archivo Excel: {str(e)}"
    print(error_msg, file=sys.stderr)
    sys.exit(1)